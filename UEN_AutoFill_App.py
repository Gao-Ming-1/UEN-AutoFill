import streamlit as st
import pandas as pd
import sqlite3
import re
import io
import os
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="UEN Autofill", page_icon="🔍", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background-color: #F7F6F2; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 2rem; max-width: 1100px; }
.element-container { margin-bottom: 0 !important; }
.stMarkdown { margin-bottom: 0 !important; }
.top-banner { background:#1A1A1A; color:#F7F6F2; padding:1.5rem 2rem; border-radius:12px; margin-bottom:1rem; }
.top-banner h1 { font-size:1.6rem; font-weight:600; margin:0; letter-spacing:-0.5px; }
.top-banner p  { font-size:0.85rem; color:#9A9A9A; margin:0.2rem 0 0 0; }
.badge { background:#3DBA6F; color:white; font-size:0.7rem; font-weight:600; padding:0.2rem 0.6rem; border-radius:20px; letter-spacing:0.5px; text-transform:uppercase; }
.pdpa-box { background:#F5F0FF; border-left:3px solid #7B5EA7; border-radius:0 8px 8px 0; padding:0.75rem 1rem; font-size:0.82rem; color:#3D2A6B; margin-bottom:1rem; line-height:1.6; }
.pdpa-box strong { color:#5A3D9A; }
.card-top { background:white; border:1px solid #E8E6DF; border-radius:12px 12px 0 0; border-bottom:none; padding:1.4rem 1.5rem 0.8rem 1.5rem; }
.card-mid { background:white; border-left:1px solid #E8E6DF; border-right:1px solid #E8E6DF; padding:0 1.5rem 1.4rem 1.5rem; }
.card-sep { background:white; border:1px solid #E8E6DF; border-bottom:none; padding:1.4rem 1.5rem 0.8rem 1.5rem; }
.card-bot { background:white; border:1px solid #E8E6DF; border-radius:0 0 12px 12px; border-top:none; padding:0 1.5rem 1.4rem 1.5rem; }
.step-label { font-size:0.72rem; font-weight:600; letter-spacing:1px; text-transform:uppercase; color:#9A9A9A; margin-bottom:0.4rem; }
.step-title { font-size:1.05rem; font-weight:600; color:#1A1A1A; margin-bottom:0; }
.stats-row { display:flex; gap:0.8rem; flex-wrap:wrap; margin-top:1rem; }
.stat-box { flex:1; min-width:120px; background:#F7F6F2; border-radius:8px; padding:0.9rem 1rem; border:1px solid #E8E6DF; text-align:center; }
.stat-number { font-family:'DM Mono',monospace; font-size:1.6rem; font-weight:500; color:#1A1A1A; }
.stat-label  { font-size:0.72rem; color:#9A9A9A; font-weight:500; margin-top:0.2rem; }
.stat-green .stat-number { color:#3DBA6F; }
.stat-orange .stat-number { color:#F5A623; }
.stat-blue .stat-number  { color:#4A90D9; }
.stat-red .stat-number   { color:#E85454; }
.cell-ref { font-family:'DM Mono',monospace; background:#F0EEE8; border:1px solid #DDD9CE; padding:0.1rem 0.5rem; border-radius:4px; font-size:0.8rem; color:#555; }
.stDownloadButton button { background:#1A1A1A !important; color:white !important; border:none !important; border-radius:8px !important; font-family:'DM Sans',sans-serif !important; font-weight:500 !important; padding:0.6rem 1.4rem !important; font-size:0.9rem !important; width:100% !important; }
.stDownloadButton button:hover { background:#333 !important; }
.stButton button { background:#3DBA6F !important; color:white !important; border:none !important; border-radius:8px !important; font-family:'DM Sans',sans-serif !important; font-weight:500 !important; padding:0.6rem 1.4rem !important; font-size:0.9rem !important; }
.stButton button:hover { background:#35A862 !important; }
.stSelectbox > div > div, .stNumberInput > div > div > input { border-radius:8px !important; border-color:#DDD9CE !important; font-family:'DM Sans',sans-serif !important; }
.info-box { background:#EBF5FF; border-left:3px solid #4A90D9; border-radius:0 8px 8px 0; padding:0.75rem 1rem; font-size:0.85rem; color:#2C5F8A; margin-bottom:0; }
.warn-box { background:#FFF8EC; border-left:3px solid #F5A623; border-radius:0 8px 8px 0; padding:0.75rem 1rem; font-size:0.85rem; color:#7A5500; margin-bottom:0; }
.dl-section { background:#F7F6F2; border:1px solid #E8E6DF; border-radius:10px; padding:1.2rem 1.4rem; margin-top:1rem; }
.dl-section-title { font-size:0.78rem; font-weight:600; letter-spacing:0.8px; text-transform:uppercase; color:#9A9A9A; margin-bottom:0.8rem; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
STOP_WORDS = {"pte","ltd","limited","private","co","corp","sdn","bhd","llp","inc","and","the","of","in","for","by","at"}
NA_PATTERNS = [
    re.compile(r'^n\.?a\.?$',re.I), re.compile(r'^n/a$',re.I),
    re.compile(r'^nil$',re.I),       re.compile(r'^self$',re.I),
    re.compile(r'^-+$'),             re.compile(r"^'+$"),
    re.compile(r"^'[-@\s]*$"),       re.compile(r'^[^a-zA-Z0-9]+$'),
]
GENERIC_NON_COMPANIES = {"freelance","freelancer","self-employed","unemployed"}

DB_PATHS = [
    "./database_1.db",
    "./database_2.db",
    "./database_3.db",
    "./database_4.db",
]

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def is_na(value: str) -> bool:
    v = value.strip()
    if not v: return False
    for p in NA_PATTERNS:
        if p.match(v): return True
    return v.lower() in GENERIC_NON_COMPANIES

def is_valid_uen(value: str) -> bool:
    v = value.strip()
    if not v or ' ' in v: return False
    if len(v) < 6 or len(v) > 15: return False
    if not re.search(r'[A-Za-z]', v): return False
    if not re.match(r'^[0-9A-Za-z]+$', v): return False
    return True

def normalise(s: str) -> str:
    return re.sub(r'\s+',' ', re.sub(r'[.\-,()&\'/\\]',' ', s.lower())).strip()

def meaningful_tokens(norm: str) -> list:
    return [t for t in norm.split() if len(t) > 1 and t not in STOP_WORDS]

def col_letter_to_index(col_str: str) -> int:
    idx = 0
    for ch in col_str.upper().strip(): idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def col_index_to_letter(idx: int) -> str:
    result, idx = "", idx + 1
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result

def clean_val(v) -> str:
    s = str(v).strip()
    return "" if s in ("nan","None","<NA>","NaN") else s

def clear_session_data():
    for key in ["processed_df","stats","preview_two_col"]:
        st.session_state.pop(key, None)


# ═══════════════════════════════════════════════════════════════════════════════
#  FTS5 SETUP
#
#  The FTS5 table indexes both company_name AND aliases so that tier-5 FTS5
#  queries can match on alias text too (e.g. a query of "IHRP" will match
#  the aliases column even when no company_name contains "IHRP").
#
#  uen is stored UNINDEXED — we only need it for retrieval, not searching.
#
#  If a previous version of the table was built without the aliases column,
#  ensure_fts5 drops and rebuilds it automatically on first startup.
# ═══════════════════════════════════════════════════════════════════════════════
def ensure_fts5(db_path: str) -> None:
    conn = sqlite3.connect(db_path)

    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='companies_fts'"
    ).fetchone()

    needs_build = False
    if row is None:
        needs_build = True
    elif row[0] and "aliases" not in row[0]:
        # Old table missing aliases column — drop and rebuild
        conn.execute("DROP TABLE IF EXISTS companies_fts")
        conn.commit()
        needs_build = True

    if needs_build:
        conn.execute("""
            CREATE VIRTUAL TABLE companies_fts
            USING fts5(
                company_name,
                aliases,
                uen UNINDEXED,
                tokenize='unicode61 remove_diacritics 1'
            )
        """)
        conn.execute("""
            INSERT INTO companies_fts(company_name, aliases, uen)
            SELECT company_name, COALESCE(aliases, ''), uen
            FROM companies
        """)
        conn.commit()
    conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
#  INDEX BUILD
#
#  Memory structures:
#    exact        dict  norm_name  -> uen    tier-1  O(1) exact name
#    alias_exact  dict  norm_alias -> uen    tier-2  O(1) exact alias  ← checked
#                                            immediately after tier-1 so short
#                                            alias-only names are never missed
#    fw_index     dict  first_word -> [(norm_name, uen, alias_norms_tuple)]
#                                            tier-3/4 substring scan
#    fts_conns    list  open read-only connections  tier-5 FTS5 fallback
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner="Loading reference database…")
def build_indexes(db_paths: tuple):
    exact       = {}
    alias_exact = {}
    fw_index    = defaultdict(list)
    fts_conns   = []
    seen        = set()

    for db_path in db_paths:
        if not os.path.exists(db_path):
            continue

        ensure_fts5(db_path)

        conn_ro = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True,
                                  check_same_thread=False)
        conn_ro.execute("PRAGMA journal_mode=OFF")
        conn_ro.execute("PRAGMA cache_size=-32000")
        fts_conns.append(conn_ro)

        rows = conn_ro.execute(
            "SELECT company_name, uen, aliases FROM companies"
        ).fetchall()

        for raw_name, uen, alias_raw in rows:
            raw_name  = (raw_name  or "").strip()
            uen       = (uen       or "").strip()
            alias_raw = (alias_raw or "").strip()
            if not raw_name and not uen:
                continue

            norm_name = sys.intern(normalise(raw_name))
            uen       = sys.intern(uen.upper())
            if not norm_name or norm_name in seen:
                continue
            seen.add(norm_name)

            # tier-1
            exact[norm_name] = uen

            # alias structures
            alias_norms = ()
            if alias_raw:
                parts = []
                for a in alias_raw.split(","):
                    a = a.strip()
                    if not a:
                        continue
                    na = sys.intern(normalise(a))
                    if na:
                        parts.append(na)
                        if na not in alias_exact:
                            alias_exact[na] = uen
                alias_norms = tuple(parts)

            # fw_index: bucket by first word + additional meaningful words
            words = norm_name.split()
            if not words:
                continue
            entry = (norm_name, uen, alias_norms)
            fw_index[words[0]].append(entry)
            for w in words[1:]:
                if w not in STOP_WORDS and len(w) > 2 and w != words[0]:
                    fw_index[w].append(entry)

    return exact, dict(fw_index), alias_exact, fts_conns, len(exact)


# ─── FTS5 LOOKUP  (tier-5 fallback) ──────────────────────────────────────────
# Searches the full FTS5 table — which indexes both company_name and aliases —
# so alias-only matches (e.g. "IHRP") are also caught here as a safety net.
def fts_lookup(norm_query: str, fts_conns: list) -> str:
    tokens = meaningful_tokens(norm_query)
    if not tokens or not fts_conns:
        return ""
    fts_q = " AND ".join(f'"{t}"' for t in tokens)
    for conn in fts_conns:
        try:
            # Querying the table name (not a column) searches ALL indexed columns
            row = conn.execute(
                "SELECT uen FROM companies_fts WHERE companies_fts MATCH ? LIMIT 1",
                (fts_q,)
            ).fetchone()
            if row and row[0]:
                return row[0].upper()
        except sqlite3.OperationalError:
            pass
    return ""


# ─── FIND UEN ─────────────────────────────────────────────────────────────────
#
# Tier 1  exact name dict              O(1)
# Tier 2  exact alias dict             O(1)   ← checked right after tier-1.
#                                              Previously this was checked only
#                                              after fw_index scanning, causing
#                                              alias-only queries (e.g. "IHRP")
#                                              to fall all the way to FTS5 and
#                                              still miss (old FTS5 had no
#                                              aliases column).  Now fixed.
# Tier 3  fw_index name substring      O(1) bucket + small list scan
# Tier 4  fw_index alias substring     same loop as tier-3 (no extra pass)
# Tier 5  FTS5 on-disk                 O(log N), searches name + aliases

def find_uen(typed_name: str,
             exact: dict, fw_index: dict, alias_exact: dict,
             fts_conns: list) -> str:

    norm_query = normalise(typed_name)
    if not norm_query:
        return ""

    # ── Tier 1: exact name ───────────────────────────────────────────────────
    if norm_query in exact:
        return exact[norm_query]

    # ── Tier 2: exact alias ──────────────────────────────────────────────────
    # Must be checked here — before fw_index — so that a query which IS an alias
    # but is NOT a substring of any official name (e.g. "IHRP") is caught at
    # O(1) cost rather than falling through all tiers and being missed.
    if norm_query in alias_exact:
        return alias_exact[norm_query]

    nq_len = len(norm_query)
    best_uen, best_score = "", -1

    # ── Tiers 3 + 4: fw_index substring scan ─────────────────────────────────
    seen_cands: dict[str, tuple] = {}
    for w in norm_query.split():
        if w in fw_index:
            for entry in fw_index[w]:
                nn = entry[0]
                if nn not in seen_cands:
                    seen_cands[nn] = entry

    for norm_name, uen, alias_norms in seen_cands.values():
        en_len = len(norm_name)

        # Tier 3: official name substring with ratio guard
        if nq_len <= en_len:
            sl, ratio, hit = nq_len, nq_len / en_len, norm_query in norm_name
        else:
            sl, ratio, hit = en_len, en_len / nq_len, norm_name in norm_query
        if hit and ratio >= 0.5:
            score = 5000 + sl
            if score > best_score:
                best_score, best_uen = score, uen
            continue  # skip alias check for this entry — name match is better

        # Tier 4: alias substring
        for alias in alias_norms:
            al_len = len(alias)
            if nq_len <= al_len:
                asl, ar, ah = nq_len, nq_len / al_len if al_len else 0, norm_query in alias
            else:
                asl, ar, ah = al_len, al_len / nq_len if nq_len else 0, alias in norm_query
            if ah and ar >= 0.5:
                score = 4000 + asl
                if score > best_score:
                    best_score, best_uen = score, uen
                break

    if best_score >= 4000:
        return best_uen

    # ── Tier 5: FTS5 on-disk ─────────────────────────────────────────────────
    fts_result = fts_lookup(norm_query, fts_conns)
    if fts_result:
        return fts_result

    return best_uen if best_score >= 0 else ""


# ─── PROCESS ──────────────────────────────────────────────────────────────────
def process_df(df, name_col_idx, uen_col_idx, header_row_idx, end_row_1based,
               exact, fw_index, alias_exact, fts_conns):

    result_df  = df.copy()
    filled = replaced = already = na_count = no_match = 0
    data_start = header_row_idx + 1
    data_end   = (end_row_1based - 1) if end_row_1based > 0 else (len(df) - 1)
    nc = df.columns[name_col_idx]
    uc = df.columns[uen_col_idx]

    _lookup_cache: dict[str, str] = {}

    def cached_find(name: str) -> str:
        key = normalise(name)
        if key not in _lookup_cache:
            _lookup_cache[key] = find_uen(name, exact, fw_index, alias_exact, fts_conns)
        return _lookup_cache[key]

    for i in range(data_start, min(data_end + 1, len(df))):
        row = df.iloc[i]
        rn  = clean_val(row.iloc[name_col_idx]) if name_col_idx < len(row) else ""
        ru  = clean_val(row.iloc[uen_col_idx])  if uen_col_idx  < len(row) else ""
        tn, tu = rn.strip(), ru.strip()

        # Both empty → NA
        if not tn and not tu:
            result_df.at[i, nc] = "NA"; result_df.at[i, uc] = "NA"
            na_count += 1; continue

        # Name empty
        if not tn:
            if is_na(tu):
                result_df.at[i, nc] = "NA"; result_df.at[i, uc] = "NA"
                na_count += 1
            continue

        # Name is a placeholder → NA
        if is_na(tn):
            result_df.at[i, nc] = "NA"; result_df.at[i, uc] = "NA"
            na_count += 1; continue

        # Name pasted into UEN column → look up
        if tn.lower() == tu.lower():
            m = cached_find(tn); result_df.at[i, uc] = m
            replaced += bool(m); no_match += not bool(m); continue

        # UEN is a placeholder → look up
        if is_na(tu):
            m = cached_find(tn); result_df.at[i, uc] = m
            filled += bool(m); no_match += not bool(m); continue

        # UEN fails format check → replace
        if tu and not is_valid_uen(tu):
            m = cached_find(tn); result_df.at[i, uc] = m
            replaced += bool(m); no_match += not bool(m); continue

        # Valid UEN present → keep (normalise to uppercase)
        if tu and is_valid_uen(tu):
            canonical = tu.upper()
            if canonical != clean_val(row.iloc[uen_col_idx]):
                result_df.at[i, uc] = canonical
            already += 1; continue

        # UEN empty → look up
        m = cached_find(tn); result_df.at[i, uc] = m
        filled += bool(m); no_match += not bool(m)

    return result_df, {"filled":filled,"replaced":replaced,"already":already,
                       "na":na_count,"no_match":no_match}


# ─── EXCEL BUILDERS ───────────────────────────────────────────────────────────
def build_full_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, header=False)
    return buf.getvalue()

def build_uen_only_excel(df, name_col_idx, uen_col_idx, header_row_idx):
    out_df = df.copy().fillna("").replace({"nan":"","None":"","<NA>":""})
    hdr_name = clean_val(out_df.iloc[header_row_idx, name_col_idx]) or "Company Name"
    hdr_uen  = clean_val(out_df.iloc[header_row_idx, uen_col_idx])  or "UEN"
    data_rows = []
    for i in range(header_row_idx + 1, len(out_df)):
        nv = clean_val(out_df.iloc[i, name_col_idx])
        uv = clean_val(out_df.iloc[i, uen_col_idx])
        if nv == "NA" and uv == "NA": continue
        data_rows.append((nv, uv))

    wb = Workbook(); ws = wb.active; ws.title = "UEN Results"
    hf    = Font(bold=True, color="F7F6F2", name="Arial", size=10)
    hfill = PatternFill("solid", fgColor="1A1A1A")
    ws.append([hdr_name, hdr_uen])
    for c in range(1, 3):
        cell = ws.cell(row=1, column=c)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    alt = PatternFill("solid", fgColor="F7F6F2")
    wht = PatternFill("solid", fgColor="FFFFFF")
    df_ = Font(name="Arial", size=10)
    uf  = Font(name="Courier New", size=10, color="1A6E3F")
    for rn, (nv, uv) in enumerate(data_rows, start=2):
        ws.append([nv, uv])
        fill = alt if rn % 2 == 0 else wht
        ws.cell(rn,1).font = df_; ws.cell(rn,1).fill = fill
        ws.cell(rn,2).font = uf;  ws.cell(rn,2).fill = fill
        ws.row_dimensions[rn].height = 18
    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        ws.column_dimensions[cl].width = min(
            max((len(str(c.value or "")) for c in col_cells), default=10) + 4, 60)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="top-banner">
  <div>
    <div style="display:flex;align-items:center;gap:0.75rem;margin-bottom:0.3rem;">
      <h1>UEN Autofill</h1><span class="badge">Singapore</span>
    </div>
    <p>Upload a spreadsheet → map your columns → download with UENs filled in</p>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="pdpa-box">
  🔒 <strong>Privacy Notice (PDPA)</strong> &nbsp;·&nbsp;
  This tool processes your uploaded file solely for UEN matching.
  Uploaded data is <strong>not stored, logged, or shared</strong> — it exists only in temporary memory
  during your session and is discarded when you close or refresh this page.
</div>
""", unsafe_allow_html=True)

missing = [p for p in DB_PATHS if not os.path.exists(p)]
if len(missing) == len(DB_PATHS):
    st.markdown('<div class="warn-box">⚠️ No database shards found. '
                'Place database_1.db … database_4.db in the app root directory.</div>',
                unsafe_allow_html=True)
    st.stop()

exact, fw_index, alias_exact, fts_conns, total_records = build_indexes(tuple(DB_PATHS))
st.markdown(f'<div class="info-box">✅ Reference database loaded — '
            f'<strong>{total_records:,}</strong> company records</div>',
            unsafe_allow_html=True)
st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

# ── STEP 1 ────────────────────────────────────────────────────────────────────
st.markdown('<div class="card-top"><div class="step-label">Step 1</div>'
            '<div class="step-title">Upload your file</div></div>', unsafe_allow_html=True)
st.markdown('<div class="card-mid">', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "file", type=["xlsx","xls","csv"],
    label_visibility="collapsed", on_change=clear_session_data)
st.markdown('</div>', unsafe_allow_html=True)

if uploaded_file is None:
    st.markdown('<div class="card-bot"><div class="info-box">'
                '👆 Upload an Excel (.xlsx / .xls) or CSV file to get started.'
                '</div></div>', unsafe_allow_html=True)
    st.stop()

@st.cache_data(show_spinner="Reading file…", max_entries=3)
def load_file(file_bytes, file_name):
    if file_name.endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str)
    return pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str, engine="openpyxl")

file_bytes = uploaded_file.read()
try:
    raw_df = load_file(file_bytes, uploaded_file.name)
except Exception as e:
    st.error(f"Could not read file: {e}"); st.stop()

num_rows, num_cols = raw_df.shape
col_letters = [col_index_to_letter(i) for i in range(num_cols)]
row_options  = [f"Row {i}" for i in range(min(num_rows, 500))]

# ── STEP 2 ────────────────────────────────────────────────────────────────────
st.markdown('<div class="card-sep"><div class="step-label">Step 2</div>'
            '<div class="step-title">Map your columns</div></div>', unsafe_allow_html=True)
st.markdown('<div class="card-mid">', unsafe_allow_html=True)
c1, c2, c3, c4 = st.columns(4)
with c1: name_col_letter = st.selectbox("Company Name column", col_letters, index=0)
with c2: uen_col_letter  = st.selectbox("UEN column", col_letters, index=min(1, len(col_letters)-1))
with c3: header_row_sel  = st.selectbox("Header row", row_options, index=0,
                               help="Row 0 = first row of file")
with c4: end_row_input   = st.number_input("Last data row (0 = auto)",
                               min_value=0, max_value=num_rows, value=0, step=1)
st.markdown('</div>', unsafe_allow_html=True)

name_col_idx   = col_letter_to_index(name_col_letter)
uen_col_idx    = col_letter_to_index(uen_col_letter)
header_row_idx = int(header_row_sel.split()[1])

# ── STEP 3: PREVIEW ───────────────────────────────────────────────────────────
st.markdown('<div class="card-sep"><div class="step-label">Step 3</div>'
            '<div class="step-title">Preview</div></div>', unsafe_allow_html=True)
st.markdown('<div class="card-mid">', unsafe_allow_html=True)

preview_df = raw_df.copy().fillna("").replace({"nan":"","None":"","<NA>":""})
preview_df.columns = [f"{col_index_to_letter(i)}  (col {i+1})" for i in range(num_cols)]
name_col_label = f"{name_col_letter}  (col {name_col_idx+1})"
uen_col_label  = f"{uen_col_letter}  (col {uen_col_idx+1})"

def highlight_cols(df):
    s = pd.DataFrame("", index=df.index, columns=df.columns)
    mask = df.index >= header_row_idx
    if name_col_label in df.columns:
        s.loc[mask, name_col_label] = "background-color:#EBF5FF;color:#2C5F8A;font-weight:500;"
    if uen_col_label in df.columns:
        s.loc[mask, uen_col_label]  = "background-color:#EBFAF2;color:#1A6E3F;font-weight:500;"
    return s

st.dataframe(preview_df.style.apply(highlight_cols, axis=None),
             width='stretch', height=min(600, 38 + num_rows * 35))

pi1, pi2 = st.columns(2)
with pi1:
    st.markdown(
        f'🔵 <span class="cell-ref">{name_col_letter}</span> Company Name &nbsp;&nbsp;'
        f'🟢 <span class="cell-ref">{uen_col_letter}</span> UEN',
        unsafe_allow_html=True)
with pi2:
    data_end_row = (end_row_input - 1) if end_row_input > 0 else (num_rows - 1)
    st.markdown(
        f'Rows to process: <span class="cell-ref">{max(0, data_end_row - header_row_idx)}</span>'
        f' &nbsp;(header = <span class="cell-ref">{header_row_sel}</span>)',
        unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ── STEP 4 ────────────────────────────────────────────────────────────────────
st.markdown('<div class="card-sep"><div class="step-label">Step 4</div>'
            '<div class="step-title">Process &amp; Download</div></div>', unsafe_allow_html=True)
st.markdown('<div class="card-bot">', unsafe_allow_html=True)

if st.button("▶  Run UEN Autofill", use_container_width=True):
    with st.spinner("Looking up UENs…"):
        processed_df, stats = process_df(
            raw_df, name_col_idx, uen_col_idx,
            header_row_idx, int(end_row_input),
            exact, fw_index, alias_exact, fts_conns)
    st.session_state["processed_df"]    = processed_df
    st.session_state["stats"]           = stats
    st.session_state["preview_two_col"] = True
    st.success("Done!")

if "processed_df" in st.session_state:
    s = st.session_state["stats"]
    st.markdown(f"""
    <div class="stats-row">
      <div class="stat-box stat-green"><div class="stat-number">{s['filled']}</div><div class="stat-label">✅ UENs filled</div></div>
      <div class="stat-box stat-blue"><div class="stat-number">{s['replaced']}</div><div class="stat-label">🔄 Replaced</div></div>
      <div class="stat-box"><div class="stat-number">{s['already']}</div><div class="stat-label">⚪ Already valid</div></div>
      <div class="stat-box stat-orange"><div class="stat-number">{s['na']}</div><div class="stat-label">🚫 Marked NA</div></div>
      <div class="stat-box stat-red"><div class="stat-number">{s['no_match']}</div><div class="stat-label">❌ No match</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    out_name = Path(uploaded_file.name).stem + "_processed"

    st.markdown('<div class="dl-section"><div class="dl-section-title">📥 Download results</div>',
                unsafe_allow_html=True)
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        st.markdown("**Full sheet** *(original columns + UEN filled)*")
        st.download_button("⬇  Excel (.xlsx)",
            build_full_excel(st.session_state["processed_df"]),
            file_name=out_name+".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with dl2:
        st.markdown("**Full sheet** *(CSV)*")
        st.download_button("⬇  CSV",
            st.session_state["processed_df"].to_csv(index=False, header=False),
            file_name=out_name+".csv", mime="text/csv", use_container_width=True)
    with dl3:
        st.markdown("**Company Name + UEN only**")
        st.download_button("⬇  UEN Results (.xlsx)",
            build_uen_only_excel(
                st.session_state["processed_df"],
                name_col_idx, uen_col_idx, header_row_idx),
            file_name=out_name+"_uen_only.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

    with st.expander("Preview processed output", expanded=True):
        two_col_view = st.toggle(
            "Show Company Name & UEN columns only",
            value=st.session_state.get("preview_two_col", True),
            key="preview_toggle")
        st.session_state["preview_two_col"] = two_col_view

        out_df = (st.session_state["processed_df"]
                  .copy().fillna("").replace({"nan":"","None":"","<NA>":""}))

        if two_col_view:
            sliced = out_df.iloc[header_row_idx:, [name_col_idx, uen_col_idx]].copy()
            new_cols = [
                str(sliced.iloc[0,0]) or f"{name_col_letter} — Company Name",
                str(sliced.iloc[0,1]) or f"{uen_col_letter} — UEN",
            ]
            sliced = sliced.iloc[1:].copy()
            sliced.columns = new_cols
            sliced.index = range(len(sliced))
            st.dataframe(sliced, width='stretch', height=400)
        else:
            full = out_df.iloc[header_row_idx:].copy()
            hv = [str(v) if str(v) not in ("","nan","None") else f"Col {col_index_to_letter(i)}"
                  for i, v in enumerate(full.iloc[0])]
            full = full.iloc[1:].copy()
            full.columns = hv
            full.index = range(len(full))
            st.dataframe(full, width='stretch', height=400)

    st.markdown(
        '<div style="margin-top:1rem;font-size:0.78rem;color:#9A9A9A;text-align:center;">'
        '🔒 Session data is held in memory only and discarded when you close or refresh this page.'
        '</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)
